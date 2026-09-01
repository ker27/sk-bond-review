#!/usr/bin/env python3
"""抽取 xlsx 全部 sheet：每格输出 坐标 = repr(值)。

用法: python3 extract_xlsx.py <输入.xlsx> <输出.txt> [maxcols=21]
- data_only=True 取公式缓存值；取到 None 说明无缓存，需用 Excel 打开另存。
- maxcols 给足 21：审计版报表常有"合并调整"列+多家子公司列，发债口径在 J/T 列。
- 先 dump 全部 坐标=值 看清列布局再取数，不要凭印象猜列。
"""
import sys
from openpyxl import load_workbook


def main(path, out, maxcols=21):
    wb = load_workbook(path, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"\n##### SHEET: {ws.title} (max_row={ws.max_row}, max_col={ws.max_column}) #####")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                max_col=min(ws.max_column, maxcols)):
            for cell in row:
                if cell.value is not None:
                    lines.append(f"{cell.coordinate} = {cell.value!r}")
    with open(out, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"sheets={len(wb.worksheets)} -> {out}")


if __name__ == '__main__':
    maxcols = int(sys.argv[3]) if len(sys.argv) > 3 else 21
    main(sys.argv[1], sys.argv[2], maxcols)
